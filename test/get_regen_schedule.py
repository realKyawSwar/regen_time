import pandas as pd
import numpy as np
from time import perf_counter
import openpyxl
import datetime


today = datetime.date.today()
start_date = datetime.date(2021, 4, 23)
delta = today - start_date
print(delta.days)
print(today)
# # path = r"R:/MAINT/PUBLIC/Public SDE/SDE Regen Status/SDE- Regen & Carrier Change Schedule.xlsm"
path = r"SDE- Regen & Carrier Change Schedule.xlsm"
# # path = "R:\MAINT\SDE\Sputter\Sputter Group\Elgin\Equipment PM.xlsx"

# wb_obj = openpyxl.load_workbook(path)

# sheet = wb_obj.active


# def getRegendate(line):
#     rowNo = 4*(line-200)+1
#     result = None
#     for value in sheet.iter_rows(min_row=5, max_row=5,
#                                  min_col=497+delta.days,
#                                  max_col=521+delta.days,
#                                  values_only=True):
#         print(value)
#     # for value in sheet.iter_rows(min_row=4, max_row=4,
#     #                              min_col=496+delta.days,
#     #                              max_col=500+delta.days,
#     #                              values_only=True):
#     #     print(value)
#         for x, i in enumerate(value):
#             if i and 'REG' in i.strip():
#                 result = x + 1
#                 break
#     return result



# def get_regen():
#     line = []
#     regen_date = []
#     for i in range(201, 212):
#         if i != 206 and i != 207:
#             x = getRegendate(i)
#             if x:
#                 line.append(i)
#                 print(i)
#                 print(datetime.timedelta(x)+today)
#                 regen_date.append(datetime.timedelta(x)+today)
#             else:
#                 pass
#         else:
#             pass
#     df = pd.DataFrame({'line': line, 'regen_date': regen_date})
#     return df


# print(get_regen())


# print(getRegendate(202))
# print(getTodaycol())

# print("{} days to Regen".format(getRegendate()))
# for value in sheet.iter_rows(min_row=5, max_row=5, min_col=8, max_col=30, values_only=True):
#     for i in value:
#         print(i)


# regen_path = r"R:/MAINT/PUBLIC/Public SDE/SDE Regen Status/SDE- Regen & Carrier Change Schedule.xlsm"
rgx = pd.ExcelFile(path)

df = pd.read_excel(rgx, sheet_name='2019-2021',
                   index_col=None, na_values=['NA'],
                   usecols="A, AGC:AGL")[2:47]
# df = df[df['EPS'].notna()]
df = df.dropna(subset=list(df.columns), how='all')

df.to_excel("regen_schedule.xlsx", index=False)


# import config

# def getTodaycol():
#     # searching
#     for value in sheet.iter_rows(min_row=4, max_row=4, min_col=490,
#                                  values_only=True):
#         for x, i in enumerate(value):
#             if i.date() == today:
#                 result = x
#                 break
#     return(result)
