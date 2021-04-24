import pandas as pd
import numpy as np
# import mus_regen
# import trouble_report
import config
# from database import Postgres
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, colors
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
# from openpyxl.drawing.fill import PatternFillProperties, ColorChoice


# df1.to_csv('Regentimingtab.csv')
df1 = pd.read_csv('Regentimingtab.csv').replace(np.nan, "")
df1 = df1.drop(['Unnamed: 0'], axis=1)
# listy = dataframe_to_rows(df1, index=False, header=True)
# for i in listy:
#     print(i)


def get_avg():
    df = pd.read_csv('pastdatatab.csv')
    return df.drop(['Unnamed: 0'], axis=1)


wb = Workbook()
# Create a few styles
bold_font = Font(bold=True)
big_red_text = Font(color=colors.RED, size=20)
center_aligned_text = Alignment(horizontal="center")


def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")
    rows = list(rows)
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(left=cell.border.left, right=cell.border.right,
                            top=cell.border.top, bottom=cell.border.bottom)
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side
            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def regen_time_tab():
    start_row = 16
    # # ws = wb.active
    df1["Target"] = 9
    ws = wb.create_sheet("Regen Timing", 0)
    ws.sheet_view.showGridLines = False
    ws.title = "Regen Timing"
    listy = dataframe_to_rows(df1, index=False, header=True)

    for row, row_entries in enumerate(listy, start=16):
        for column, value in enumerate(row_entries, start=1):
            ws.cell(column=column, row=row, value=value)
    # for r in dataframe_to_rows(df1, index=False, header=True):
    #     ws.append(r)
    for cell in ws[start_row]:
        cell.font = bold_font
        cell.alignment = center_aligned_text
        cell.fill = PatternFill(start_color="feedc6",
                                end_color="feedc6",
                                fill_type="solid")
    # border
    BORDER_LIST = [f'A{start_row}:H{ws.max_row}', f'A{start_row}:H{start_row}']
    for pos in BORDER_LIST:
        set_border(ws, pos)
    # center characters
    for row_cells in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                  min_col=1, max_col=2):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center')
    for row_cells in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                  min_col=3):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='right')
    ws.column_dimensions['G'].width = 11.5
    chart1 = BarChart()
    data = Reference(worksheet=ws, min_row=start_row+1, max_row=ws.max_row,
                     min_col=3, max_col=3)
    cats = Reference(ws, min_row=start_row+1, max_row=ws.max_row,
                     min_col=1, max_col=2)
    # set the title of the chart
    chart1.title = f'{config.month_name} Regen Timing'
    chart1.add_data(data, titles_from_data=False)
    chart1.legend = None
    chart1.set_categories(cats)
    # Style the lines
    s1 = chart1.series[0]
    s1.graphicalProperties.solidFill = "ff4400"
    target_line = LineChart()
    target = Reference(worksheet=ws, min_row=start_row+1, max_row=ws.max_row,
                       min_col=8)
    target_line.add_data(target)
    chart1 += target_line
    ws.add_chart(chart1, "J1")
    print('done')


def avg_tab():
    df = get_avg()
    df['date'] = pd.to_datetime(df['date'])
    new_row = {'date': 'Target', 'timing': 9, 'numbers_of_regen': ''}
    df['date'] = df['date'].dt.strftime('%b-%y')
    # append row to the dataframe
    df = df.append(new_row, ignore_index=True)
    ws2 = wb.create_sheet("Past Data", 1)
    ws2.title = "Past Data"
    ws2.sheet_view.showGridLines = False
    # ws.title = "Regen Timing"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)
        # border
    BORDER_LIST = [f'A1:C{ws2.max_row}', 'A1:C1']
    for pos in BORDER_LIST:
        set_border(ws2, pos)
    for cell in ws2[1]:
        cell.font = bold_font
        cell.alignment = center_aligned_text
        cell.fill = PatternFill(start_color="feedc6", end_color="feedc6",
                                fill_type="solid")
    ws2.column_dimensions['A'].width = 11.5
    ws2.column_dimensions['C'].width = 17
    # ws = wb.active


def avg_chart():
    chart1 = BarChart()
    ws = wb["Past Data"]
    data = Reference(worksheet=ws, min_row=1, max_row=ws.max_row,
                     min_col=2, max_col=2)
    cats = Reference(worksheet=ws, min_row=2, max_row=ws.max_row,
                     min_col=1, max_col=1)
    # set the title of the chart
    chart1.title = f'Average Regen Timing'
    chart1.add_data(data, titles_from_data=True)
    chart1.legend = None
    chart1.set_categories(cats)
    count_line = LineChart()
    count = Reference(worksheet=ws, min_row=2, max_row=ws.max_row-1, min_col=3)
    count_line.add_data(count)
    chart1 += count_line
    # set data labels and styles
    s1 = chart1.series[0]
    s1.dLbls = DataLabelList()
    s1.dLbls.showVal = True
    pt = DataPoint(idx=ws.max_row-3)
    pt.graphicalProperties.solidFill = "ff9900"
    pt1 = DataPoint(idx=ws.max_row-2)
    pt1.graphicalProperties.solidFill = "00ff00"
    s1.dPt.append(pt)
    s1.dPt.append(pt1)
    # print(type(ws.max_row))
    wb["Regen Timing"].add_chart(chart1, "A1")


regen_time_tab()
avg_tab()
avg_chart()
wb.save(f'{config.month_name} Regen Timing1.xlsx')
