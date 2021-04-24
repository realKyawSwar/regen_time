import pandas as pd
import numpy as np
import os
from modules import config, trouble_report, mus_regen, conversion
from modules.database import Postgres
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
# import trouble_report
# import config
# import mus_regen

# \\hdsingapore.showadenkohd.com\SHDS\MAINT\PUBLIC\Lim KS\Weekly and Monthly MU Monitoring\Monthly MBO\

wb = Workbook()
merged = pd.merge(mus_regen.regen_time(), trouble_report.regen(),
                  on=['Date', 'Line'], how='left').replace(np.nan, "")
# merged.to_excel('lol.xlsx')
merged['Duration'] = np.where(merged['DURATION']-merged['CHECK'] < 20,
                              merged['DURATION'], merged['CHECK'])
merged['Hours'] = merged.Duration.apply(lambda x: (x/60)).round(decimals=2)
clean = merged[['Date', 'Line', 'Hours', 'Target', 'Program', 'Disk Type',
                'Duration', 'Activities', 'Leakages']]


def newDir():
    main_dir = r"//hdsingapore.showadenkohd.com/SHDS/MAINT/PUBLIC/Lim KS/Weekly and Monthly MU Monitoring/Monthly MBO"
    dirName = main_dir + "/" + str(config.year) + "/" + str(config.month_name)
    # Create target directory & all intermediate directories if don't exists
    try:
        os.makedirs(dirName)
        print("Directory ", dirName,  " Created ")
    except FileExistsError:
        # print("Directory ", dirName,  " already exists")
        pass
    # Create target directory & all intermediate directories if don't exists
    if not os.path.exists(dirName):
        os.makedirs(dirName)
    else:
        print("Directory Okay")
    return dirName


def report_tab():
    ws1 = wb.create_sheet("MUS Raw", 2)
    ws1.sheet_view.showGridLines = False
    ws1.title = "MUS Raw"
    for r in dataframe_to_rows(merged, index=False, header=True):
        ws1.append(r)


def get_avg():
    # Calculate avg and upload to DB and retrieve 12 month data
    # return df
    listo = config.get_last_month_dates()
    avg = round(clean["Hours"].mean(), 2)
    count = clean["Hours"].count()
    listy = [listo[0], avg, count]
    myPg = None
    try:
        myPg = Postgres("128.53.1.198/5432", "spt_db", "spt_admin",
                        "sptadmin")
        df1 = myPg.query(f"SELECT * FROM regen.timing where date ="
                         f" '{listo[0].replace('/', '-')}' ")[0]
        if df1.empty:
            # print(df1)
            myPg.connect()
            print("Connection succeeded..")
            # ['2016-08-01,16.5,6', '2016-09-01,15.71,7']
            vals = [",".join(str(bit) for bit in listy)]
            for i in vals:
                y = i.split(",")
                x = tuple(y)
                strSQL = f"INSERT INTO regen.timing values {x}"
                myPg.execute(strSQL)
            myPg.commit()
        else:
            print('exist!')
    except Exception as err:
        print(f"Upload Error Occured: {err}")
    finally:
        df = myPg.query("SELECT * FROM regen.timing "
                        "WHERE date >= date_trunc('month', now()) - interval '13 month' and"
                        " date < date_trunc('month', now())")[0]
        # df.to_csv('pastdatatab.csv')
        if myPg is not None:
            myPg.close()
    return df


def set_border(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")
    rows = list(rows)
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(left=cell.border.left, right=cell.border.right,
                            top=cell.border.top, bottom=cell.border.bottom)
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side
            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border


def regen_time_tab():
    youtec_df = trouble_report.youtec()
    df1 = pd.merge(clean, youtec_df, on='Date', how='left').replace(np.nan, "")
    df1 = df1[['Date', 'Line', 'Hours', 'Target', 'Program', 'Disk Type',
               'Duration', 'CVD Change', 'Activities', 'Leakages']]
    start_row = 16
    # # ws = wb.active
    ws = wb.create_sheet("Regen Timing", 0)
    ws.freeze_panes = 'A17'
    ws.sheet_view.showGridLines = False
    ws.title = "Regen Timing"
    listy = dataframe_to_rows(df1, index=False, header=True)

    for row, row_entries in enumerate(listy, start=16):
        for column, value in enumerate(row_entries, start=1):
            ws.cell(column=column, row=row, value=value)
    # for r in dataframe_to_rows(df1, index=False, header=True):
    #     ws.append(r)
    for cell in ws[start_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="feedc6",
                                end_color="feedc6",
                                fill_type="solid")
    # merge
    # for i in range(start_row, ws.max_row+1):
    #     ws.merge_cells(start_row=i, start_column=9, end_row=i,
    #                    end_column=16)
    # border
    BORDER_LIST = [f'A{x}:J{x}' for x in range(start_row, ws.max_row+1)]
    for pos in BORDER_LIST:
        set_border(ws, pos)
    # center characters
    for row_cells in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                  min_col=1, max_col=2):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for row_cells in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                  min_col=3, max_col=7):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='right', vertical='center')
    for row_cells in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                  min_col=8, max_col=8):
        for cell in row_cells:
            cell.alignment = Alignment(horizontal='center',
                                       vertical='center')
    for row_cells in ws.iter_rows(min_row=start_row, max_row=ws.max_row,
                                  min_col=9, max_col=10):
        for cell in row_cells:
            cell.alignment = Alignment(vertical='center', wrapText=True)
            # cell.style.alignment.wrap_text = True
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 41
    ws.column_dimensions['J'].width = 41
    chart1 = BarChart()
    data = Reference(worksheet=ws, min_row=start_row+1, max_row=ws.max_row,
                     min_col=3, max_col=3)
    cats = Reference(ws, min_row=start_row+1, max_row=ws.max_row,
                     min_col=1, max_col=2)
    # set the title of the chart
    chart1.title = f'{config.month_name} Regen Timing'
    chart1.add_data(data, titles_from_data=False)
    chart1.legend = None
    chart1.set_categories(cats)
    # Style the lines
    s1 = chart1.series[0]
    s1.graphicalProperties.solidFill = "ff4400"
    target_line = LineChart()
    target = Reference(worksheet=ws, min_row=start_row+1, max_row=ws.max_row,
                       min_col=4)
    target_line.add_data(target)
    chart1 += target_line
    ws.add_chart(chart1, "I1")
    print('done')


def avg_tab():
    df = get_avg()
    df['date'] = pd.to_datetime(df['date'])
    new_row = {'date': 'Target', 'timing': 9, 'numbers_of_regen': ''}
    df['date'] = df['date'].dt.strftime('%b-%y')
    # append row to the dataframe
    df = df.append(new_row, ignore_index=True)
    ws2 = wb.create_sheet("Past Data", 1)
    ws2.title = "Past Data"
    ws2.sheet_view.showGridLines = False
    # ws.title = "Regen Timing"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws2.append(r)
        # border
    BORDER_LIST = [f'A1:C{ws2.max_row}', 'A1:C1']
    for pos in BORDER_LIST:
        set_border(ws2, pos)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="feedc6", end_color="feedc6",
                                fill_type="solid")
    ws2.column_dimensions['A'].width = 11.5
    ws2.column_dimensions['C'].width = 17
    # ws = wb.active


def avg_chart():
    chart1 = BarChart()
    ws = wb["Past Data"]
    data = Reference(worksheet=ws, min_row=1, max_row=ws.max_row,
                     min_col=2, max_col=2)
    cats = Reference(worksheet=ws, min_row=2, max_row=ws.max_row,
                     min_col=1, max_col=1)
    # set the title of the chart
    chart1.title = f'Average Regen Timing'
    chart1.add_data(data, titles_from_data=True)
    chart1.legend = None
    chart1.set_categories(cats)
    count_line = LineChart()
    count = Reference(worksheet=ws, min_row=2, max_row=ws.max_row-1, min_col=3)
    count_line.add_data(count)
    chart1 += count_line
    # set data labels and styles
    s1 = chart1.series[0]
    s1.dLbls = DataLabelList()
    s1.dLbls.showVal = True
    pt = DataPoint(idx=ws.max_row-3)
    pt.graphicalProperties.solidFill = "ff9900"
    pt1 = DataPoint(idx=ws.max_row-2)
    pt1.graphicalProperties.solidFill = "00ff00"
    s1.dPt.append(pt)
    s1.dPt.append(pt1)
    # print(type(ws.max_row))
    wb["Regen Timing"].add_chart(chart1, "A1")


def conversion_tab():
    ws1 = wb.create_sheet("conversion", 4)
    ws1.sheet_view.showGridLines = False
    ws1.title = "conversion"
    df = conversion.report_convo()
    for r in dataframe_to_rows(df, index=False, header=True):
        ws1.append(r)


regen_time_tab()
avg_tab()
avg_chart()
report_tab()
conversion_tab()

# wb.save(f'{newDir()}/{config.month_name} Regen Timing.xlsx')
wb.save('test.xlsx')
